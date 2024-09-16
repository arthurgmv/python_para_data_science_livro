from openpyxl import Workbook

planilha = Workbook()

aba1 = planilha.create_sheet('Aba01', 0)
aba2 = planilha.create_sheet('Aba02')

for sheet in planilha:
    print(sheet.title)

del planilha['Sheet']

print('-'*20)

for sheet in planilha:
    print(sheet.title)

planilha.save('planilha.xlsx')